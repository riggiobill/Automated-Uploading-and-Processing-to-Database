
#~ python Upload.py

print("Initializing Upload.py...")
print("Importing libraries...")

from openpyxl import Workbook
import logging
import time
from suds.client import Client
import datetime
from openpyxl import load_workbook
import os.path
from PasswordHolder import PasswordHolder
from PasswordHolder import Email_pwd 

print("Library importing successful.")

#* These are the filenames for the raw Tracker Asia and MAWB Bags files provided.
#* These variables can be changed to whatever the recurring file name for the raw spreadsheets will be.

print("Loading target spreadsheets...")

TApath = os.path.dirname(__file__) + "\\HKG BRU AWB TRACKER 02 JUNE 2022.xlsx"
TAworkbook = load_workbook(TApath,data_only=True)

MAWBpath = os.path.dirname(__file__) + "\\MAWB-Bags (24).xlsx"
MAWBworkbook = load_workbook(MAWBpath)

print("Spreadsheets loaded successfully.")

#&-----------------------------------------------------
#&      - Makes all the changes needed
#&-----------------------------------------------------

print("Preparing spreadsheets for upload...")

#^ Tracker Asia

#* Sets the title of the main sheet to Tracker Asia.
TrackerAsia_sheet = TAworkbook["TRACKER AIR_RAIL"]
TrackerAsia_sheet.title = "Tracker Asia"

#* Sets the two sheets to be removed as variables, then removes them.
Rail_sheet = TAworkbook["RAIL"]
Reporting_sheet = TAworkbook["REPORTING "]

TAworkbook.remove(Rail_sheet)
TAworkbook.remove(Reporting_sheet)

print("Pre-upload transformations successful.")

#~ ---- 

print("Renaming spreadsheets...")

#^ MAWB
#* Sets the title of the MAWB main sheet to MAWB-Bags.
MAWB_Bags_sheet = MAWBworkbook["Sheet0"]
MAWB_Bags_sheet.title = "MAWB-Bags"

print("Spreadsheets successfully renamed.")





#&------------------------------------------------------------
#&      - Defines the send_email function for error reporting
#&------------------------------------------------------------

def send_email(user, pwd, recipient, subject, body):
    import smtplib

    FROM = user
    TO = recipient if isinstance(recipient, list) else [recipient]
    SUBJECT = subject
    TEXT = body

    # Prepare actual message
    message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    """ % (FROM, ", ".join(TO), SUBJECT, TEXT)
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(user, pwd)
        server.sendmail(FROM, TO, message)
        server.close()
        print('successfully sent the mail')
    except:
        print("failed to send mail")

##* Email login info : can be edited and updated based on client needs. Will need email login access info to pass to the smtp client.

user1 = 'riggiobill@gmail.com'
pwd1 = Email_pwd 
recipient1 = 'briggio@praestoconsulting.com'
subject1 = "BPost Automatic Upload - Error Reporting"
body1 = "This email would be sent if there is an error of some kind."

#&--------------------------------------------------------
#&      - Checks all the changes have been made correctly
#&--------------------------------------------------------


print("Confirming changes made to spreadsheets...")

#* Checks to see that the worksheet title has been correctly changed.
TAcheck_title = TrackerAsia_sheet.title

if TAcheck_title != "Tracker Asia":
    print("Code not executed properly : closing program.")

    body1 = "Spreadsheet title change unsuccessful. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()
#~ ---
MAWBcheck_title = MAWB_Bags_sheet.title

if MAWBcheck_title != "MAWB-Bags":
    print("Code not executed properly : closing program.")

    body1 = "Spreadsheet title change unsuccessful. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#* Checks to see that the other sheets have been sucessfully removed.
check_sheets = str(TAworkbook.sheetnames)

if check_sheets != "['Tracker Asia']":
    print("Code not executed properly : closing program.")

    body1 = "Spreadsheet sheet removal unsuccessful. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#* Saves and renames the workbook.
print("Now saving...")
TAworkbook.save(filename=os.path.dirname(__file__) + "//Tracker Asia.xlsx")
MAWBworkbook.save(filename=os.path.dirname(__file__) + "//MAWB-Bags.xlsx")

#* Checks to see that the save and rename was successful.
if os.path.isfile('Tracker Asia.xlsx'):
    holder = 0
else:
    print("Issue renaming the file, code exiting")

    body1 = "Spreadsheet renaming and saving unsuccessful. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()
#~ ---
if os.path.isfile('MAWB-Bags.xlsx'):
    holder = 0
else:
    print("Issue renaming the file, code exiting")

    body1 = "Spreadsheet renaming and saving unsuccessful. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

print("Changes successfully confirmed.")


#&-----------------------------------------------------------------
#&      - Check all the columns and data types before Birst upload
#&-----------------------------------------------------------------


print("Double checking column names and data types...")

#* using Column Name as a value for comparison, and cell.data_type for assessing data type 
#* Problem arises when theres a cell with a lot of empty values, case in point Deposit/CMR/HAWB which has many blanks.
    #* The blanks come up as data type n (presumably for none) as opposed to s (for string)

#^ Tracker Asia

#* MAWB
if(TrackerAsia_sheet["B1"].value != "MAWB"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,2).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Carrier 
if(TrackerAsia_sheet["C1"].value != "CARRIER"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,3).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Destination 
if(TrackerAsia_sheet["D1"].value != "DESTINATION"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,4).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Shipper 
if(TrackerAsia_sheet["E1"].value != "SHIPPER"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,5).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Product
if(TrackerAsia_sheet["F1"].value != "PRODUCT"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,6).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Clearance Mode
if(TrackerAsia_sheet["G1"].value != "CLEARANCE MODE"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,7).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Bags
if(TrackerAsia_sheet["H1"].value != "BAGS"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,8).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* WT KG
if(TrackerAsia_sheet["I1"].value != "WT (KG)"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,9).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#! Is supposed to be an integer, only has NA in the text. Inquire.
"""
#* DD WT
if(TrackerAsia_sheet["J1"].value != "DD WT"):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,10).data_type != "n"):
    print("Bad data detected; check columns and try again.")
    exit()
"""
#! Is supposed to be an integer, only has NA in the text. Inquire.

#~ ---

#* MF BAG QTY
if(TrackerAsia_sheet["K1"].value != "MF BAG QTY"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,11).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* MF WT
if(TrackerAsia_sheet["L1"].value != "MF WT"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,12).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* PK WT
if(TrackerAsia_sheet["M1"].value != "PK WT"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,13).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#! Is supposed to be an integer, only has NA in the text. Inquire.
"""
#* MF V DD
if(TrackerAsia_sheet["N1"].value != "MF V DD"):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,14).data_type != "n"):
    print("Bad data detected; check columns and try again.")
    exit()
"""
#! Is supposed to be an integer, only has NA in the text. Inquire.

#~ ---

#* MAWB DATE
if(str(TrackerAsia_sheet["O1"].value) != "MAWB DATE "):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,15).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* AIRPORT
if(str(TrackerAsia_sheet["P1"].value) != "AIRPORT"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,16).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#! 
"""
print(TrackerAsia_sheet["Q1"].value)
print(TrackerAsia_sheet.cell(2,17).value)
print(TrackerAsia_sheet.cell(2,17).data_type)

#* WEEK
if(str(TrackerAsia_sheet["Q1"].value) != "WEEK"):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,17).data_type != "n"):
    print("Bad data detected; check columns and try again.")
    exit()
""" 
#! 

#~ ---

#! 
"""
print(TrackerAsia_sheet["R1"].value)
print(TrackerAsia_sheet.cell(2,18).value)
print(TrackerAsia_sheet.cell(2,18).data_type)

#* MONTH
if(str(TrackerAsia_sheet["R1"].value) != "MONTH"):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,18).data_type != "n"):
    print("Bad data detected; check columns and try again.")
    exit()
""" 
#! 

#~ ---

#* DEP. DATE & TIME
if(str(TrackerAsia_sheet["S1"].value) != "DEP. DATE & TIME"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,19).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* BRU/LGG ARRIVAL DATE/TIME
if(str(TrackerAsia_sheet["T1"].value) != '''BRU/LGG
ARRIVAL DATE/TIME'''):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,20).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* # BAGS ARRIVED
if(str(TrackerAsia_sheet["U1"].value) != "# BAGS ARRIVED"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,21).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---
"""
#! 
print(TrackerAsia_sheet["V1"].value)
print(TrackerAsia_sheet.cell(2,22).value)
print(TrackerAsia_sheet.cell(2,22).data_type)

#* Release Note Available  Date/Time
if(str(TrackerAsia_sheet["V1"].value) != "Release Note Available  Date/Time"):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,22).data_type != "d"):
    print("Bad data detected; check columns and try again.")
    exit()
#!
"""
#~ ---

#*  LGG-F4U-LLI/ATF HANDOVER DATE/TIME
if(str(TrackerAsia_sheet["W1"].value) != """ LGG-F4U-LLI/ATF
HANDOVER DATE/TIME"""):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,23).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#*  PROCESS @ EMC/LX/ JOSTLGG/ AX/NBX
if(str(TrackerAsia_sheet["X1"].value) != "PROCESS @ EMC/LX/ JOSTLGG/ AX/NBX"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,24).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#*  EMC or LX or JOSTLGG DELIVERY DATE/TIME
if(str(TrackerAsia_sheet["Y1"].value) != """EMC or LX or JOSTLGG
DELIVERY DATE/TIME"""):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,25).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#*  BAGS
if(str(TrackerAsia_sheet["Z1"].value) != "BAGS"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,26).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* WT (kg)
if(str(TrackerAsia_sheet["AA1"].value) != "WT (kg)"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,27).data_type != "n"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* SPLIT
if(str(TrackerAsia_sheet["AB1"].value) != "SPLIT "):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,28).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* CUSTOMS EMC/LCA  F4U/4PX  YUN
if(str(TrackerAsia_sheet["AC1"].value) != "CUSTOMS EMC/LCA  F4U/4PX  YUN"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,29).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* CUSTOMS GAD  EMC/4PX/F4U/YUN/UBI ETC... CLEARANCE DATE/TIME 
if(str(TrackerAsia_sheet["AD1"].value) != """CUSTOMS GAD  EMC/4PX/F4U/YUN/UBI ETC...
CLEARANCE DATE/TIME """):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,30).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---
"""
#!
print(TrackerAsia_sheet["AE1"].value)
print(TrackerAsia_sheet.cell(2,31).value)
print(TrackerAsia_sheet.cell(2,31).data_type)

#* BRU/LGG ARRIVAL DATE/TIME 1
if(str(TrackerAsia_sheet["AE1"].value) != '''BRU/LGG
ARRIVAL DATE/TIME '''):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,31).data_type != "d"):
    print("Bad data detected; check columns and try again.")
    exit()
#!
"""
#!  This issue follows for the light blue section, from BRU/LGG ARRIVAL DATE/TIME to COMMENTS.
#!  These columns are mostly blank, and so are returning cell type n for None.
#!  Cannot be confirmed through Excel if left blank - will resolve the columns I can and address this later if there's time.

#^ # BAGS ARRIVED/ NOTIFIED

#^ EMC / LX DELIVERY DATE/TIME

#^ BAGS 2

#^ WT (kg) 2

#^ CUSTOMS CLEARANCE DATE/TIME

#^ EMC / LX DELIVERY DATE/TIME 1

#^ BAGS 3

#^ BAGS TO BE DLVD

#^ COMMENTS

#~ ---


#& Following this, the columns are all derived measures, amounts of time. May have to be reconstructed in Birst as derived values.

#* HK/CN - BRU
if(str(TrackerAsia_sheet["AO1"].value) != "HK/CN - BRU"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,41).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* BRU - CC
if(str(TrackerAsia_sheet["AP1"].value) != "BRU - CC"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,42).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* BRU - EMC
if(str(TrackerAsia_sheet["AQ1"].value) != "BRU - EMC"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,43).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* BRU - F4U
if(str(TrackerAsia_sheet["AR1"].value) != "BRU - F4U"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,44).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---
"""
#!
print(TrackerAsia_sheet["AS1"].value)
print(TrackerAsia_sheet.cell(2,45).value)
print(TrackerAsia_sheet.cell(2,45).data_type)

#* F4U - CC
if(str(TrackerAsia_sheet["AS1"].value) != "F4U - CC"):
    print("Bad data detected; check columns and try again.")
    exit()

if(TrackerAsia_sheet.cell(2,45).data_type != "d"):
    print("Bad data detected; check columns and try again.")
    exit()
#!
"""

#~ ---

#* CC - EMC
if(str(TrackerAsia_sheet["AT1"].value) != "CC - EMC"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,46).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* HK/CN - CC
if(str(TrackerAsia_sheet["AU1"].value) != "HK/CN - CC"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,47).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* HK/CN - EMC
if(str(TrackerAsia_sheet["AV1"].value) != "HK/CN - EMC"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(TrackerAsia_sheet.cell(2,48).data_type != "d"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ----------

#^ MAWB

#* BagNumber
if(str(MAWB_Bags_sheet["A1"].value) != "BagNumber"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(MAWB_Bags_sheet.cell(2,1).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* ReceiverCountry
if(str(MAWB_Bags_sheet["B1"].value) != "ReceiverCountry"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(MAWB_Bags_sheet.cell(2,2).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* MAWB_Number
if(str(MAWB_Bags_sheet["C1"].value) != "MAWB_Number"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(MAWB_Bags_sheet.cell(2,3).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* MAWB_date
#! Is a string in the format of a date. Birst should be able to pick this up regardless, but if its a problem can use an iter.rows to convert each cell in this column to a date.


if(str(MAWB_Bags_sheet["D1"].value) != "MAWB_date"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(MAWB_Bags_sheet.cell(2,4).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

#~ ---

#* Transport

if(str(MAWB_Bags_sheet["E1"].value) != "Transport"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()

if(MAWB_Bags_sheet.cell(2,5).data_type != "s"):
    print("Bad data detected; check columns and try again.")

    body1 = "Bad column name or data type detected. Program closed."
    send_email(user1,pwd1,recipient1,subject1,body1)

    exit()


#~ --------

print("Column names and data types successfully confirmed.")


#&------------------------------------------------------
#&      - Prompt a workflow to run from Birst and 
#&          import and process the new files.
#&------------------------------------------------------

#* Space details: update as required
#* This section will currently not run - the raw and cleaned files must be accessible through a Birst Agent 
#* in order to automate a workflow for import and processing. As such, the following has been commented out but
#* left in for possible future use.

spaceID = '7c2c9d72-e83d-4db6-a337-d00cd16b2040'
spacelogin = 'briggio@praestoconsulting.com'
spacepassword = PasswordHolder
workflowname = 'BIS DEV Automated Upload'


class workflowClass:
    """Class to run workflow"""

    def __init__(self, wsdl_url, space_id, user_name, password):
        self.wsdl_url = wsdl_url
        self.space_id = space_id
        self.user_name = user_name
        self.client = None
        self.token = None
        self.password = password
        self.connected = False

    def create_client(self):
        """Creates the client"""
        # Create client from BWS wsdl
        logging.info('Creating WSDL client...')
        self.client = Client(self.wsdl_url)

    def login(self):
        """Log in to the Birst Web Service"""
        # Log into BWS and get initial token
        logging.info('Login to Birst Web Service...')
        self.token = self.client.service.Login(self.user_name, self.password)
        self.connected = True

    def logout(self):
        """Log out of the Birst Web Service"""
        self.client.service.Logout(self.token)
        self.connected = False

    # Execute a workflow
    def run_workflow(self, WorkflowName, retry_count, retry_pause):
        """Run named workflow with options for number of retries if logged out, and pause between retries"""

        # Check if the workflow is already running
        print('Check that the workflow is not currently running')
        workflow_status = self.client.service.getLatestWorkflowExecutionStatus(self.token, WorkflowName)
        print('Pre execution workflow status is: ' + workflow_status)

        if workflow_status == 'RUNNING':
            logging.info('Workflow is already running, ETL has failed')
            return 1

        # Run the workflow
        workflow = self.client.service.executeWorkflow(self.token, WorkflowName)
        logging.info(workflow)

        # Initialise retry counter to 0
        self.retry_counter = 0

        # Poll for workflow to begin
        workflow_status = self.client.service.getLatestWorkflowExecutionStatus(self.token, WorkflowName)

        while workflow_status != 'RUNNING':
            logging.info('Waiting for workflow to start')
            workflow_status = self.client.service.getLatestWorkflowExecutionStatus(self.token, WorkflowName)
            time.sleep(1)

        logging.info('Workflow has Started')

        # Poll for completion and return status
        status = self.poll_workflow(WorkflowName, retry_count, retry_pause)
        return status

    # Check the status of a running workflow
    def poll_workflow(self, WorkflowName, retry_count, retry_pause):
        """Poll the workflow for completion with retry logic if connection is lost"""

        # Poll status
        workflow_status = self.client.service.getLatestWorkflowExecutionStatus(self.token, WorkflowName)

        while workflow_status == 'RUNNING':

            try:
                workflow_status = self.client.service.getLatestWorkflowExecutionStatus(self.token, WorkflowName)
                logging.info('Obtaining workflow status: ' + workflow_status)

                if workflow_status != 'RUNNING':
                    logging.info('Exiting with Status: ' + workflow_status)
                    break
                else:
                    logging.info('Workflow not complete. Sleeping...')
                    time.sleep(30)

            except Exception as gen_exception:
                if self.retry_counter <= retry_count:
                    logging.info('An unexpected error has occurred:')
                    logging.info(gen_exception)
                    logging.info('Will sleep for ' + str(retry_pause) + 'secs before attempting to log in.')
                    time.sleep(retry_pause)

                    # login to Birst again and retry polling
                    self.login()
                    self.retry_counter = self.retry_counter + 1
                    print("Retry polling workflow, time = " + str(datetime.datetime.today()))
                    self.poll_workflow(WorkflowName, retry_count, retry_pause)

                else:
                    logging.info('Workflow failed after ' + str(retry_count) + ' retries')
                    return 1

        if workflow_status != 'COMPLETED':
            return 1
        else:
            return 0


## Part of script that runs the workflow


# run workflow
x = workflowClass('https://login.bws.birst.com/CommandWebservice.asmx?wsdl', spaceID, spacelogin, spacepassword)
x.create_client()
x.login()
#*
x.run_workflow(workflowname, 50, 30)
#*
x.logout()
print("Finished running workflow at :" + str(datetime.datetime.today()))


print("Finished running Upload.py")